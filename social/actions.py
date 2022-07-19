from django.core.exceptions import PermissionDenied
from django.http import HttpResponse
from openpyxl import Workbook
from io import StringIO
from .models import Post
from django.contrib.auth.models import User


def export_as_xls(modeladmin, request, queryset):
    """
    Generic xls export admin action.
    """
    if not request.user.is_staff:
        raise PermissionDenied
    
    opts = modeladmin.model._meta

    wb = Workbook()

    ws = wb.create_sheet('%d' % export_as_xls.ws_no)

    col = 1
    field_names = ['username','name','bio']
    # field_names += [User._meta.get_field('username')]
    # for field in opts.get_fields():
    #     if field != User and field != 'dp':
    #         field_names += [field]

    export_as_xls.ws_no += 1
    # write header row
    for col in range(1,4):
        c = ws.cell(row = 1, column = col)
        c.value = field_names[col-1]

    r = 2
    q = queryset.values_list('user','name','bio')
    # Write data rows
    for obj in q:
        col = 1
        row = [str(obj[0]),str(obj[1]),str(obj[2])]
        for field in field_names:
                val = row[col-1]
                c = ws.cell(row = r, column = col)
                c.value = val
                print(c.value)
                col = col + 1
        r = r + 1
    s = wb.get_sheet_by_name('Sheet')
    wb.remove_sheet(s)
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=%s.xls' % str(opts).replace('.', '_')
    wb.save(response)
    return response

export_as_xls.short_description = "Export selected objects to XLS"
export_as_xls.ws_no = 0